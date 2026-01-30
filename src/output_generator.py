# Output Generation for Cell Scheduling Engine.
# Version: 1.0.0
# Generates reports, Gantt charts, and exports in various formats.

import json
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Literal

from .constants import CellColor
from .calculated_fields import (
    PRIORITY_PAST_DUE, PRIORITY_TODAY, PRIORITY_EXPEDITE, PRIORITY_FUTURE,
    SCHED_CLASS_A, SCHED_CLASS_B, SCHED_CLASS_C, SCHED_CLASS_D, SCHED_CLASS_E
)
from .multi_cell_scheduler import MultiCellScheduleResult
from .scheduler import CellScheduleResult, ScheduledPanel
from .method_variants import SchedulingMethod, SchedulingVariant
from .method_evaluation import MethodEvaluation, evaluate_result


PRIORITY_NAMES = {
    PRIORITY_PAST_DUE: "Past Due",
    PRIORITY_TODAY: "Today",
    PRIORITY_EXPEDITE: "Expedite",
    PRIORITY_FUTURE: "Future"
}


def generate_schedule_report(
    result: MultiCellScheduleResult,
    method: SchedulingMethod | None = None,
    variant: SchedulingVariant | None = None,
    include_details: bool = True
) -> str:
    """Generate a comprehensive text report for a schedule.
    
    Args:
        result: MultiCellScheduleResult to report on.
        method: Optional scheduling method used.
        variant: Optional variant used.
        include_details: Whether to include panel-level details.
    
    Returns:
        Multi-line report string.
    """
    lines = []
    
    # Header
    lines.append("=" * 80)
    lines.append("CELL SCHEDULING ENGINE - SCHEDULE REPORT")
    lines.append("=" * 80)
    lines.append("")
    
    # Schedule info
    lines.append(f"Schedule Date: {result.schedule_date}")
    lines.append(f"Shift Duration: {result.shift_minutes} minutes")
    if method:
        method_names = {
            SchedulingMethod.PRIORITY_FIRST: "Priority First",
            SchedulingMethod.MINIMUM_FORCED_IDLE: "Minimum Forced Idle",
            SchedulingMethod.MAXIMUM_OUTPUT: "Maximum Output",
            SchedulingMethod.MOST_RESTRICTED_MIX: "Most Restricted Mix"
        }
        variant_names = {
            SchedulingVariant.JOB_FIRST: "Job First",
            SchedulingVariant.TABLE_FIRST: "Table First"
        }
        lines.append(f"Method: {method_names.get(method, str(method))}")
        if variant:
            lines.append(f"Variant: {variant_names.get(variant, str(variant))}")
    lines.append(f"Status: {result.status}")
    lines.append("")
    
    # Summary
    lines.append("-" * 80)
    lines.append("SUMMARY")
    lines.append("-" * 80)
    lines.append(f"Total Panels Scheduled: {result.total_panels}")
    lines.append(f"Total Operator Minutes: {result.total_operator_minutes}")
    lines.append(f"Jobs Scheduled: {len(result.job_assignments)}")
    lines.append(f"Jobs Unscheduled: {len(result.unscheduled_jobs)}")
    lines.append("")
    
    # Cell breakdown
    lines.append("-" * 80)
    lines.append("CELL BREAKDOWN")
    lines.append("-" * 80)
    
    for cell_color in sorted(result.cell_results.keys()):
        cr = result.cell_results[cell_color]
        lines.append(f"\n{cell_color} CELL:")
        lines.append(f"  Status: {cr.status}")
        lines.append(f"  Total Panels: {cr.total_panels}")
        lines.append(f"  Table 1 Panels: {len(cr.table1_panels)}")
        lines.append(f"  Table 2 Panels: {len(cr.table2_panels)}")
        lines.append(f"  Operator Time: {cr.total_operator_time} minutes")
        lines.append(f"  Forced Operator Idle: {cr.forced_operator_idle} minutes")
        
        for table_id, idle in cr.forced_table_idle.items():
            lines.append(f"  {table_id} Idle: {idle} minutes")
        
        # End-of-day prep panels
        if cr.table1_prep:
            lines.append(f"  T1 PREP FOR TOMORROW: {cr.table1_prep.job_id} (LAYOUT ends {cr.table1_prep.end_time})")
        if cr.table2_prep:
            lines.append(f"  T2 PREP FOR TOMORROW: {cr.table2_prep.job_id} (LAYOUT ends {cr.table2_prep.end_time})")
    
    lines.append("")
    
    # Job assignments
    lines.append("-" * 80)
    lines.append("JOB ASSIGNMENTS")
    lines.append("-" * 80)
    
    for assignment in result.job_assignments:
        job = assignment.job
        calc = assignment.calc
        lines.append(f"\nJob: {job.job_id}")
        lines.append(f"  Description: {job.description[:60]}...")
        lines.append(f"  Cell: {assignment.cell_color}, Table: {assignment.table_num}")
        lines.append(f"  Panels: {assignment.panels_to_schedule}")
        lines.append(f"  Priority: {calc.priority} ({PRIORITY_NAMES.get(calc.priority, 'Unknown')})")
        lines.append(f"  SCHED_CLASS: {calc.sched_class}")
        lines.append(f"  BUILD_DATE: {calc.build_date}")
        lines.append(f"  Molds: {job.molds} × {job.mold_type}")
        if assignment.is_on_table_today:
            lines.append(f"  ON_TABLE_TODAY: Yes (starts_with_pour={assignment.starts_with_pour})")
    
    # Unscheduled jobs
    if result.unscheduled_jobs:
        lines.append("")
        lines.append("-" * 80)
        lines.append("UNSCHEDULED JOBS")
        lines.append("-" * 80)
        
        for item in result.unscheduled_jobs:
            job = item[0]
            calc = item[1] if len(item) > 2 else None
            reason = item[-1]
            
            lines.append(f"\nJob: {job.job_id}")
            if calc:
                lines.append(f"  Req By: {job.req_by}, Build Date: {calc.build_date}, Priority: {calc.priority}")
            lines.append(f"  Reason: {reason}")
    
    # Detailed panel schedule
    if include_details:
        lines.append("")
        lines.append("-" * 80)
        lines.append("DETAILED PANEL SCHEDULE")
        lines.append("-" * 80)
        
        for cell_color in sorted(result.cell_results.keys()):
            cr = result.cell_results[cell_color]
            
            lines.append(f"\n{cell_color} CELL:")
            
            for table_name, panels in [("Table 1", cr.table1_panels), ("Table 2", cr.table2_panels)]:
                if panels:
                    lines.append(f"\n  {table_name}:")
                    for panel in panels:
                        lines.append(f"    Panel {panel.panel_index}: Job {panel.job_id}")
                        for task_name in ["SETUP", "LAYOUT", "POUR", "CURE", "UNLOAD"]:
                            task = panel.tasks.get(task_name)
                            if task:
                                op_marker = "*" if task.requires_operator else " "
                                lines.append(
                                    f"      {op_marker}{task_name}: {task.start_time}-{task.end_time} "
                                    f"({task.duration} min)"
                                )
    
    # Warnings
    if result.warnings:
        lines.append("")
        lines.append("-" * 80)
        lines.append("WARNINGS")
        lines.append("-" * 80)
        for warning in result.warnings:
            lines.append(f"  - {warning}")
    
    lines.append("")
    lines.append("=" * 80)
    lines.append(f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 80)
    
    return "\n".join(lines)


def generate_gantt_text(
    result: MultiCellScheduleResult,
    width: int = 100
) -> str:
    """Generate ASCII Gantt chart for all cells.
    
    Args:
        result: MultiCellScheduleResult to visualize.
        width: Character width of the chart.
    
    Returns:
        ASCII Gantt chart string.
    """
    lines = []
    lines.append("=" * width)
    lines.append("GANTT CHART - ALL CELLS")
    lines.append("=" * width)
    
    # Time scale
    shift = result.shift_minutes
    scale = (width - 20) / shift  # Leave room for labels
    
    # Time axis
    time_axis = " " * 20
    for t in range(0, shift + 1, 60):
        pos = int(t * scale)
        marker = f"{t:3d}"
        if pos + len(marker) <= width - 20:
            time_axis = time_axis[:pos] + marker + time_axis[pos + len(marker):]
    lines.append(time_axis)
    lines.append(" " * 20 + "|" + "-" * int(shift * scale))
    
    # Task symbols
    TASK_CHARS = {
        "SETUP": "S",
        "LAYOUT": "L",
        "POUR": "P",
        "CURE": ".",
        "UNLOAD": "U"
    }
    
    for cell_color in sorted(result.cell_results.keys()):
        cr = result.cell_results[cell_color]
        
        lines.append(f"\n{cell_color}:")
        
        for table_name, panels in [("_1", cr.table1_panels), ("_2", cr.table2_panels)]:
            table_id = f"{cell_color}{table_name}"
            
            # Build timeline
            timeline = [" "] * int(shift * scale + 1)
            
            for panel in panels:
                for task_name, task in panel.tasks.items():
                    char = TASK_CHARS.get(task_name, "?")
                    start_pos = int(task.start_time * scale)
                    end_pos = int(task.end_time * scale)
                    for i in range(start_pos, min(end_pos, len(timeline))):
                        timeline[i] = char
            
            lines.append(f"  {table_id:12s} |{''.join(timeline)}")
        
        # Operator timeline
        op_timeline = [" "] * int(shift * scale + 1)
        
        for panels in [cr.table1_panels, cr.table2_panels]:
            for panel in panels:
                for task_name, task in panel.tasks.items():
                    if task.requires_operator:
                        char = TASK_CHARS.get(task_name, "?")
                        start_pos = int(task.start_time * scale)
                        end_pos = int(task.end_time * scale)
                        for i in range(start_pos, min(end_pos, len(op_timeline))):
                            op_timeline[i] = char
        
        lines.append(f"  {'OPERATOR':12s} |{''.join(op_timeline)}")
    
    lines.append("")
    lines.append("Legend: S=SETUP, L=LAYOUT, P=POUR, .=CURE, U=UNLOAD")
    lines.append("=" * width)
    
    return "\n".join(lines)


def generate_html_gantt(
    result: MultiCellScheduleResult,
    title: str = "Cell Schedule"
) -> str:
    """Generate an HTML Gantt chart for all cells.
    
    Args:
        result: MultiCellScheduleResult to visualize.
        title: Title for the HTML page.
    
    Returns:
        Complete HTML document string.
    """
    shift = result.shift_minutes
    
    # Color scheme for tasks
    TASK_COLORS = {
        "SETUP": "#FF6B6B",
        "LAYOUT": "#4ECDC4",
        "POUR": "#45B7D1",
        "CURE": "#96CEB4",
        "UNLOAD": "#FFEAA7"
    }
    
    # Prep panel colors (same but with stripe pattern)
    PREP_COLORS = {
        "SETUP": "#FF6B6B",
        "LAYOUT": "#4ECDC4"
    }
    
    html = f"""<!DOCTYPE html>
<html>
<head>
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        h1 {{ color: #333; }}
        .info {{ background: white; padding: 15px; border-radius: 8px; margin-bottom: 20px; }}
        .gantt-container {{ background: white; padding: 20px; border-radius: 8px; overflow-x: auto; }}
        .cell-section {{ margin-bottom: 30px; }}
        .cell-title {{ font-size: 18px; font-weight: bold; color: #333; margin-bottom: 10px; }}
        .timeline {{ position: relative; height: 30px; background: #eee; margin: 5px 0; border-radius: 4px; }}
        .timeline-label {{ position: absolute; left: -100px; width: 90px; text-align: right; font-size: 12px; line-height: 30px; }}
        .task {{ position: absolute; height: 24px; top: 3px; border-radius: 3px; font-size: 10px; color: white; text-align: center; line-height: 24px; overflow: hidden; }}
        .prep-task {{ position: absolute; height: 24px; top: 3px; border-radius: 3px; font-size: 10px; color: white; text-align: center; line-height: 24px; overflow: hidden; border: 2px dashed #333; box-sizing: border-box; }}
        .time-axis {{ position: relative; height: 20px; margin-left: 100px; }}
        .time-marker {{ position: absolute; font-size: 10px; color: #666; }}
        .legend {{ margin-top: 20px; }}
        .legend-item {{ display: inline-block; margin-right: 20px; }}
        .legend-color {{ display: inline-block; width: 20px; height: 12px; margin-right: 5px; vertical-align: middle; border-radius: 2px; }}
        .prep-note {{ font-size: 12px; color: #666; margin-top: 10px; font-style: italic; }}
        .summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; }}
        .summary-card {{ background: #fff; border: 1px solid #ddd; border-radius: 8px; padding: 15px; }}
        .summary-card h3 {{ margin: 0 0 10px 0; font-size: 14px; color: #666; }}
        .summary-card .value {{ font-size: 24px; font-weight: bold; color: #333; }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    
    <div class="info">
        <div class="summary">
            <div class="summary-card">
                <h3>Schedule Date</h3>
                <div class="value">{result.schedule_date}</div>
            </div>
            <div class="summary-card">
                <h3>Total Panels</h3>
                <div class="value">{result.total_panels}</div>
            </div>
            <div class="summary-card">
                <h3>Jobs Scheduled</h3>
                <div class="value">{len(result.job_assignments)}</div>
            </div>
            <div class="summary-card">
                <h3>Status</h3>
                <div class="value">{result.status}</div>
            </div>
        </div>
    </div>
    
    <div class="gantt-container">
        <div class="time-axis">
"""
    
    # Time markers
    pixels_per_min = 2
    for t in range(0, shift + 1, 60):
        html += f'            <span class="time-marker" style="left: {t * pixels_per_min}px">{t}</span>\n'
    
    html += "        </div>\n"
    
    # Each cell
    for cell_color in sorted(result.cell_results.keys()):
        cr = result.cell_results[cell_color]
        
        html += f"""
        <div class="cell-section">
            <div class="cell-title">{cell_color} CELL ({cr.total_panels} panels)</div>
"""
        
        for table_name, panels in [("Table 1", cr.table1_panels), ("Table 2", cr.table2_panels)]:
            table_id = f"{cell_color}_{table_name[-1]}"
            prep_panel = cr.table1_prep if table_name == "Table 1" else cr.table2_prep
            
            html += f"""
            <div class="timeline" style="margin-left: 100px; width: {shift * pixels_per_min}px;">
                <span class="timeline-label">{table_id}</span>
"""
            
            for panel in panels:
                for task_name, task in panel.tasks.items():
                    if task.duration > 0:
                        left = task.start_time * pixels_per_min
                        width = max(task.duration * pixels_per_min, 2)
                        color = TASK_COLORS.get(task_name, "#999")
                        label = task_name[0] if width > 15 else ""
                        html += f'                <div class="task" style="left: {left}px; width: {width}px; background: {color};" title="{task_name}: {task.start_time}-{task.end_time}">{label}</div>\n'
            
            # Add prep panel if exists (dashed border style)
            if prep_panel:
                # SETUP task
                if prep_panel.setup_task.duration > 0:
                    left = prep_panel.setup_task.start_time * pixels_per_min
                    width = max(prep_panel.setup_task.duration * pixels_per_min, 2)
                    html += f'                <div class="prep-task" style="left: {left}px; width: {width}px; background: {PREP_COLORS["SETUP"]};" title="PREP SETUP: {prep_panel.setup_task.start_time}-{prep_panel.setup_task.end_time} (tomorrow)">S</div>\n'
                
                # LAYOUT task
                left = prep_panel.layout_task.start_time * pixels_per_min
                width = max(prep_panel.layout_task.duration * pixels_per_min, 2)
                html += f'                <div class="prep-task" style="left: {left}px; width: {width}px; background: {PREP_COLORS["LAYOUT"]};" title="PREP LAYOUT: {prep_panel.layout_task.start_time}-{prep_panel.layout_task.end_time} (tomorrow)">L</div>\n'
            
            html += "            </div>\n"
        
        html += "        </div>\n"
    
    # Legend
    html += """
        <div class="legend">
            <strong>Legend:</strong>
"""
    for task_name, color in TASK_COLORS.items():
        html += f'            <span class="legend-item"><span class="legend-color" style="background: {color};"></span>{task_name}</span>\n'
    
    # Count prep panels
    prep_count = sum(1 for cr in result.cell_results.values() if cr.table1_prep) + \
                 sum(1 for cr in result.cell_results.values() if cr.table2_prep)
    
    html += """
        </div>
"""
    
    if prep_count > 0:
        html += f"""        <div class="prep-note">
            <strong>Note:</strong> {prep_count} end-of-day prep panel(s) shown with dashed borders - SETUP+LAYOUT done, ready for POUR tomorrow
        </div>
"""
    
    html += """    </div>
</body>
</html>
"""
    
    return html


def export_to_json(
    result: MultiCellScheduleResult,
    method: SchedulingMethod | None = None,
    variant: SchedulingVariant | None = None,
    pretty: bool = True
) -> str:
    """Export schedule to JSON format.
    
    Args:
        result: MultiCellScheduleResult to export.
        method: Optional scheduling method used.
        variant: Optional variant used.
        pretty: Whether to format with indentation.
    
    Returns:
        JSON string.
    """
    data = {
        "schedule_date": str(result.schedule_date),
        "shift_minutes": result.shift_minutes,
        "status": result.status,
        "total_panels": result.total_panels,
        "total_operator_minutes": result.total_operator_minutes,
        "method": method.name if method else None,
        "variant": variant.name if variant else None,
        "cells": {},
        "job_assignments": [],
        "unscheduled_jobs": []
    }
    
    # Cell results
    for cell_color, cr in result.cell_results.items():
        data["cells"][cell_color] = {
            "status": cr.status,
            "total_panels": cr.total_panels,
            "total_operator_time": cr.total_operator_time,
            "forced_operator_idle": cr.forced_operator_idle,
            "forced_table_idle": cr.forced_table_idle,
            "table1_panels": [
                {
                    "panel_index": p.panel_index,
                    "job_id": p.job_id,
                    "tasks": {
                        name: {
                            "start": t.start_time,
                            "end": t.end_time,
                            "duration": t.duration,
                            "requires_operator": t.requires_operator
                        }
                        for name, t in p.tasks.items()
                    }
                }
                for p in cr.table1_panels
            ],
            "table2_panels": [
                {
                    "panel_index": p.panel_index,
                    "job_id": p.job_id,
                    "tasks": {
                        name: {
                            "start": t.start_time,
                            "end": t.end_time,
                            "duration": t.duration,
                            "requires_operator": t.requires_operator
                        }
                        for name, t in p.tasks.items()
                    }
                }
                for p in cr.table2_panels
            ],
            # End-of-day prep panels (SETUP+LAYOUT done, ready for POUR tomorrow)
            "table1_prep": {
                "job_id": cr.table1_prep.job_id,
                "setup_end": cr.table1_prep.setup_task.end_time,
                "layout_end": cr.table1_prep.layout_task.end_time
            } if cr.table1_prep else None,
            "table2_prep": {
                "job_id": cr.table2_prep.job_id,
                "setup_end": cr.table2_prep.setup_task.end_time,
                "layout_end": cr.table2_prep.layout_task.end_time
            } if cr.table2_prep else None
        }
    
    # Job assignments
    for a in result.job_assignments:
        data["job_assignments"].append({
            "job_id": a.job.job_id,
            "cell_color": a.cell_color,
            "table_num": a.table_num,
            "panels": a.panels_to_schedule,
            "priority": a.calc.priority,
            "sched_class": a.calc.sched_class,
            "build_date": str(a.calc.build_date),
            "is_on_table_today": a.is_on_table_today
        })
    
    # Unscheduled
    for item in result.unscheduled_jobs:
        job = item[0]
        calc = item[1] if len(item) > 2 else None
        reason = item[-1]
        
        unsched_entry = {
            "job_id": job.job_id,
            "reason": reason,
            "req_by": str(job.req_by) if job.req_by else None
        }
        
        if calc:
            unsched_entry["build_date"] = str(calc.build_date) if calc.build_date else None
            unsched_entry["priority"] = calc.priority
            unsched_entry["sched_class"] = calc.sched_class
        
        data["unscheduled_jobs"].append(unsched_entry)
    
    indent = 2 if pretty else None
    return json.dumps(data, indent=indent)


def generate_comparison_report(
    results: dict[tuple[SchedulingMethod, SchedulingVariant], MultiCellScheduleResult]
) -> str:
    """Generate a comparison report for multiple method results.
    
    Args:
        results: Dict mapping (method, variant) to result.
    
    Returns:
        Multi-line comparison report.
    """
    lines = []
    lines.append("=" * 100)
    lines.append("METHOD COMPARISON REPORT")
    lines.append("=" * 100)
    lines.append("")
    
    # Evaluate all
    evaluations = []
    for key in results:
        method, variant = key[0], key[1]
        result = results[key]
        eval = evaluate_result(result, method, variant)
        evaluations.append((key, eval, result))
    
    # Summary table
    lines.append("SUMMARY")
    lines.append("-" * 100)
    header = f"{'Method':<40} {'Status':<12} {'Panels':<8} {'Jobs':<6} {'Table Idle':<12} {'Op Idle':<10}"
    lines.append(header)
    lines.append("-" * 100)
    
    for (method, variant), eval, result in evaluations:
        method_name = f"{method.name} ({variant.name})"
        lines.append(
            f"{method_name:<40} "
            f"{eval.status:<12} "
            f"{eval.total_panels:<8} "
            f"{eval.total_jobs_scheduled:<6} "
            f"{eval.efficiency.forced_table_idle:<12} "
            f"{eval.efficiency.forced_operator_idle:<10}"
        )
    
    lines.append("")
    
    # Best performers
    lines.append("BEST PERFORMERS")
    lines.append("-" * 100)
    
    # Most panels
    best_panels = max(evaluations, key=lambda x: x[1].total_panels)
    lines.append(f"Most Panels: {best_panels[0][0].name} ({best_panels[0][1].name}) - {best_panels[1].total_panels} panels")
    
    # Most efficient
    best_eff = min(evaluations, key=lambda x: x[1].efficiency.forced_table_idle + x[1].efficiency.forced_operator_idle)
    total_idle = best_eff[1].efficiency.forced_table_idle + best_eff[1].efficiency.forced_operator_idle
    lines.append(f"Most Efficient: {best_eff[0][0].name} ({best_eff[0][1].name}) - {total_idle} total idle minutes")
    
    lines.append("")
    
    # Priority breakdown
    lines.append("PRIORITY BREAKDOWN")
    lines.append("-" * 100)
    
    for (method, variant), eval, result in evaluations:
        method_name = f"{method.name} ({variant.name})"
        lines.append(f"\n{method_name}:")
        for priority, metrics in sorted(eval.priority_metrics.items()):
            pname = PRIORITY_NAMES.get(priority, f"Priority {priority}")
            lines.append(f"  {pname}: {metrics.scheduled} jobs, {metrics.panels_scheduled} panels")
    
    lines.append("")
    
    # Class breakdown
    lines.append("SCHED_CLASS BREAKDOWN")
    lines.append("-" * 100)
    header = f"{'Method':<40} {'A':<6} {'B':<6} {'C':<6} {'D':<6} {'E':<6}"
    lines.append(header)
    lines.append("-" * 100)
    
    for (method, variant), eval, result in evaluations:
        method_name = f"{method.name} ({variant.name})"
        cm = eval.class_metrics
        lines.append(
            f"{method_name:<40} "
            f"{cm.class_a:<6} "
            f"{cm.class_b:<6} "
            f"{cm.class_c:<6} "
            f"{cm.class_d:<6} "
            f"{cm.class_e:<6}"
        )
    
    lines.append("")
    lines.append("=" * 100)
    lines.append(f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 100)
    
    return "\n".join(lines)


def save_all_outputs(
    result: MultiCellScheduleResult,
    output_dir: str | Path,
    method: SchedulingMethod | None = None,
    variant: SchedulingVariant | None = None,
    prefix: str = "schedule"
) -> dict[str, Path]:
    """Save all output formats to files.
    
    Args:
        result: MultiCellScheduleResult to output.
        output_dir: Directory to save files.
        method: Optional scheduling method.
        variant: Optional variant.
        prefix: Filename prefix.
    
    Returns:
        Dict of format name to file path.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    saved = {}
    
    # Text report
    report_path = output_dir / f"{prefix}_report.txt"
    with open(report_path, "w") as f:
        f.write(generate_schedule_report(result, method, variant))
    saved["report"] = report_path
    
    # Gantt text
    gantt_path = output_dir / f"{prefix}_gantt.txt"
    with open(gantt_path, "w") as f:
        f.write(generate_gantt_text(result))
    saved["gantt_text"] = gantt_path
    
    # HTML Gantt
    html_path = output_dir / f"{prefix}_gantt.html"
    with open(html_path, "w") as f:
        f.write(generate_html_gantt(result, f"Schedule - {result.schedule_date}"))
    saved["gantt_html"] = html_path
    
    return saved


def generate_cell_html_report(
    result: MultiCellScheduleResult,
    cell_color: str,
    title: str = None
) -> str:
    """Generate HTML report for a single cell with gantt chart.
    
    Args:
        result: Full schedule result.
        cell_color: Cell to generate report for (RED, BLUE, etc.)
        title: Optional title override.
    
    Returns:
        HTML string with gantt chart and cell details.
    """
    if cell_color not in result.cell_results:
        return f"<html><body><h1>No results for {cell_color}</h1></body></html>"
    
    cr = result.cell_results[cell_color]
    title = title or f"{cell_color} Cell Schedule - {result.schedule_date}"
    
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; }}
        .summary {{ background: #f8f9fa; padding: 15px; border-radius: 8px; margin-bottom: 20px; }}
        .summary-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; }}
        .stat {{ text-align: center; }}
        .stat-value {{ font-size: 28px; font-weight: bold; color: #2c3e50; }}
        .stat-label {{ font-size: 12px; color: #666; text-transform: uppercase; }}
        .gantt-container {{ margin: 20px 0; }}
        .gantt-row {{ display: flex; align-items: center; margin: 5px 0; }}
        .gantt-label {{ width: 100px; font-weight: bold; font-size: 14px; }}
        .gantt-bar {{ flex: 1; height: 30px; background: #ecf0f1; position: relative; border-radius: 3px; }}
        .task {{ position: absolute; height: 100%; border-radius: 2px; display: flex; align-items: center; justify-content: center; font-size: 10px; color: white; font-weight: bold; }}
        .task-SETUP {{ background: #e74c3c; }}
        .task-LAYOUT {{ background: #27ae60; }}
        .task-POUR {{ background: #3498db; }}
        .task-CURE {{ background: #f39c12; }}
        .task-UNLOAD {{ background: #9b59b6; }}
        .time-axis {{ display: flex; margin-left: 100px; border-top: 1px solid #ddd; padding-top: 5px; }}
        .time-marker {{ font-size: 10px; color: #888; position: absolute; }}
        .legend {{ display: flex; gap: 20px; margin: 20px 0; }}
        .legend-item {{ display: flex; align-items: center; gap: 5px; font-size: 12px; }}
        .legend-color {{ width: 20px; height: 12px; border-radius: 2px; }}
        .job-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        .job-table th, .job-table td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        .job-table th {{ background: #3498db; color: white; }}
        .job-table tr:nth-child(even) {{ background: #f8f9fa; }}
        .idle-info {{ background: #fff3e0; padding: 10px; border-radius: 5px; margin: 10px 0; }}
        @media print {{ body {{ margin: 0; }} }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    
    <div class="summary">
        <div class="summary-grid">
            <div class="stat">
                <div class="stat-value">{cr.total_panels}</div>
                <div class="stat-label">Total Panels</div>
            </div>
            <div class="stat">
                <div class="stat-value">{len(cr.table1_panels)}</div>
                <div class="stat-label">{cell_color}_1 Panels</div>
            </div>
            <div class="stat">
                <div class="stat-value">{len(cr.table2_panels)}</div>
                <div class="stat-label">{cell_color}_2 Panels</div>
            </div>
            <div class="stat">
                <div class="stat-value">{cr.forced_operator_idle}</div>
                <div class="stat-label">Operator Idle (min)</div>
            </div>
        </div>
    </div>
    
    <div class="legend">
        <div class="legend-item"><div class="legend-color" style="background:#e74c3c"></div>SETUP</div>
        <div class="legend-item"><div class="legend-color" style="background:#27ae60"></div>LAYOUT</div>
        <div class="legend-item"><div class="legend-color" style="background:#3498db"></div>POUR</div>
        <div class="legend-item"><div class="legend-color" style="background:#f39c12"></div>CURE</div>
        <div class="legend-item"><div class="legend-color" style="background:#9b59b6"></div>UNLOAD</div>
    </div>
"""
    
    # Generate Gantt chart
    shift_minutes = result.shift_minutes
    px_per_min = 1.2
    total_width = shift_minutes * px_per_min
    
    html += f"""
    <h2>Gantt Chart</h2>
    <div class="gantt-container">
        <div style="position:relative;margin-left:100px;width:{total_width}px;height:20px;margin-bottom:5px;">
"""
    
    # Time markers
    for t in range(0, shift_minutes + 1, 60):
        html += f'<span class="time-marker" style="left:{t * px_per_min}px">{t}</span>'
    
    html += "</div>"
    
    # Table 1
    html += f"""
        <div class="gantt-row">
            <div class="gantt-label">{cell_color}_1</div>
            <div class="gantt-bar" style="width:{total_width}px">
"""
    
    for panel in cr.table1_panels:
        for task_name, task in panel.tasks.items():
            if task.duration > 0:
                left = task.start_time * px_per_min
                width = max(task.duration * px_per_min, 2)
                label = task_name[0] if width > 15 else ""
                html += f'<div class="task task-{task_name}" style="left:{left}px;width:{width}px">{label}</div>'
    
    html += """
            </div>
        </div>
"""
    
    # Table 2
    html += f"""
        <div class="gantt-row">
            <div class="gantt-label">{cell_color}_2</div>
            <div class="gantt-bar" style="width:{total_width}px">
"""
    
    for panel in cr.table2_panels:
        for task_name, task in panel.tasks.items():
            if task.duration > 0:
                left = task.start_time * px_per_min
                width = max(task.duration * px_per_min, 2)
                label = task_name[0] if width > 15 else ""
                html += f'<div class="task task-{task_name}" style="left:{left}px;width:{width}px">{label}</div>'
    
    html += """
            </div>
        </div>
    </div>
"""
    
    # Idle time info
    total_table_idle = sum(cr.forced_table_idle.values()) if cr.forced_table_idle else 0
    html += f"""
    <div class="idle-info">
        <strong>Efficiency Metrics:</strong>
        Forced Operator Idle: {cr.forced_operator_idle} min |
        Forced Table Idle: {total_table_idle} min
    </div>
"""
    
    # Job details table
    html += """
    <h2>Panel Details</h2>
    <table class="job-table">
        <tr>
            <th>Table</th>
            <th>Panel</th>
            <th>Job ID</th>
            <th>SETUP</th>
            <th>LAYOUT</th>
            <th>POUR</th>
            <th>CURE</th>
            <th>UNLOAD</th>
            <th>End Time</th>
        </tr>
"""
    
    for panel in cr.table1_panels:
        setup = panel.tasks.get("SETUP")
        layout = panel.tasks.get("LAYOUT")
        pour = panel.tasks.get("POUR")
        cure = panel.tasks.get("CURE")
        unload = panel.tasks.get("UNLOAD")
        
        html += f"""
        <tr>
            <td>{cell_color}_1</td>
            <td>{panel.panel_index + 1}</td>
            <td>{panel.job_id}</td>
            <td>{setup.start_time}-{setup.end_time} ({setup.duration}m)</td>
            <td>{layout.start_time}-{layout.end_time} ({layout.duration}m)</td>
            <td>{pour.start_time}-{pour.end_time} ({pour.duration}m)</td>
            <td>{cure.start_time}-{cure.end_time} ({cure.duration}m)</td>
            <td>{unload.start_time}-{unload.end_time} ({unload.duration}m)</td>
            <td>{panel.end_time}</td>
        </tr>
"""
    
    for panel in cr.table2_panels:
        setup = panel.tasks.get("SETUP")
        layout = panel.tasks.get("LAYOUT")
        pour = panel.tasks.get("POUR")
        cure = panel.tasks.get("CURE")
        unload = panel.tasks.get("UNLOAD")
        
        html += f"""
        <tr>
            <td>{cell_color}_2</td>
            <td>{panel.panel_index + 1}</td>
            <td>{panel.job_id}</td>
            <td>{setup.start_time}-{setup.end_time} ({setup.duration}m)</td>
            <td>{layout.start_time}-{layout.end_time} ({layout.duration}m)</td>
            <td>{pour.start_time}-{pour.end_time} ({pour.duration}m)</td>
            <td>{cure.start_time}-{cure.end_time} ({cure.duration}m)</td>
            <td>{unload.start_time}-{unload.end_time} ({unload.duration}m)</td>
            <td>{panel.end_time}</td>
        </tr>
"""
    
    html += """
    </table>
    
    <div style="margin-top: 30px; font-size: 11px; color: #888; text-align: center;">
        Generated by Cell Scheduling Engine
    </div>
</body>
</html>
"""
    
    return html


def generate_cell_pdf(
    result: MultiCellScheduleResult,
    cell_color: str,
    output_path: str | Path
) -> Path:
    """Generate PDF report for a single cell using reportlab.
    
    Args:
        result: Full schedule result.
        cell_color: Cell to generate report for.
        output_path: Path to save PDF.
    
    Returns:
        Path to generated PDF.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics import renderPDF
    
    if cell_color not in result.cell_results:
        raise ValueError(f"No results for {cell_color}")
    
    cr = result.cell_results[cell_color]
    output_path = Path(output_path)
    
    # Create PDF document
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(letter),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=12
    )
    
    elements = []
    
    # Title
    elements.append(Paragraph(f"{cell_color} Cell Schedule - {result.schedule_date}", title_style))
    elements.append(Spacer(1, 12))
    
    # Summary stats
    total_table_idle = sum(cr.forced_table_idle.values()) if cr.forced_table_idle else 0
    summary_data = [
        ["Total Panels", f"{cell_color}_1 Panels", f"{cell_color}_2 Panels", "Operator Idle", "Table Idle"],
        [str(cr.total_panels), str(len(cr.table1_panels)), str(len(cr.table2_panels)), 
         f"{cr.forced_operator_idle} min", f"{total_table_idle} min"]
    ]
    summary_table = Table(summary_data, colWidths=[1.8*inch]*5)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ecf0f1')),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Panel details table
    elements.append(Paragraph("Panel Details", styles['Heading2']))
    elements.append(Spacer(1, 8))
    
    panel_headers = ["Table", "Panel", "Job ID", "SETUP", "LAYOUT", "POUR", "CURE", "UNLOAD", "End"]
    panel_data = [panel_headers]
    
    for panel in cr.table1_panels:
        setup = panel.tasks.get("SETUP")
        layout = panel.tasks.get("LAYOUT")
        pour = panel.tasks.get("POUR")
        cure = panel.tasks.get("CURE")
        unload = panel.tasks.get("UNLOAD")
        
        panel_data.append([
            f"{cell_color}_1",
            str(panel.panel_index + 1),
            panel.job_id[:15],
            f"{setup.start_time}-{setup.end_time}" if setup else "-",
            f"{layout.start_time}-{layout.end_time}" if layout else "-",
            f"{pour.start_time}-{pour.end_time}" if pour else "-",
            f"{cure.start_time}-{cure.end_time}" if cure else "-",
            f"{unload.start_time}-{unload.end_time}" if unload else "-",
            str(panel.end_time)
        ])
    
    for panel in cr.table2_panels:
        setup = panel.tasks.get("SETUP")
        layout = panel.tasks.get("LAYOUT")
        pour = panel.tasks.get("POUR")
        cure = panel.tasks.get("CURE")
        unload = panel.tasks.get("UNLOAD")
        
        panel_data.append([
            f"{cell_color}_2",
            str(panel.panel_index + 1),
            panel.job_id[:15],
            f"{setup.start_time}-{setup.end_time}" if setup else "-",
            f"{layout.start_time}-{layout.end_time}" if layout else "-",
            f"{pour.start_time}-{pour.end_time}" if pour else "-",
            f"{cure.start_time}-{cure.end_time}" if cure else "-",
            f"{unload.start_time}-{unload.end_time}" if unload else "-",
            str(panel.end_time)
        ])
    
    col_widths = [0.8*inch, 0.5*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.5*inch]
    panel_table = Table(panel_data, colWidths=col_widths)
    panel_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(panel_table)
    
    # Footer
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f"Generated by Cell Scheduling Engine v1.0.0 - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray)
    ))
    
    doc.build(elements)
    return output_path


def generate_summary_pdf(
    result: MultiCellScheduleResult,
    method_name: str,
    variant_name: str,
    output_path: str | Path
) -> Path:
    """Generate summary PDF report for entire schedule.
    
    Args:
        result: Full schedule result.
        method_name: Scheduling method used.
        variant_name: Variant used.
        output_path: Path to save PDF.
    
    Returns:
        Path to generated PDF.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    output_path = Path(output_path)
    
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=20, spaceAfter=20)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, spaceBefore=15, spaceAfter=10)
    
    elements = []
    
    # Title
    elements.append(Paragraph(f"Schedule Report - {result.schedule_date}", title_style))
    
    # Method info
    elements.append(Paragraph(f"Method: {method_name} ({variant_name})", styles['Normal']))
    elements.append(Paragraph(f"Shift: {result.shift_minutes} minutes", styles['Normal']))
    elements.append(Spacer(1, 15))
    
    # Overall summary
    elements.append(Paragraph("Summary", section_style))
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Panels", str(result.total_panels)],
        ["Jobs Scheduled", str(len(result.job_assignments))],
        ["Jobs Unscheduled", str(len(result.unscheduled_jobs))],
        ["Active Cells", str(len(result.cell_results))],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Cell breakdown
    elements.append(Paragraph("Cell Breakdown", section_style))
    
    cell_headers = ["Cell", "Total Panels", "Table 1", "Table 2", "Op Idle", "T1 Prep", "T2 Prep"]
    cell_data = [cell_headers]
    
    for cell_color, cr in result.cell_results.items():
        t1_prep = cr.table1_prep.job_id if cr.table1_prep else "-"
        t2_prep = cr.table2_prep.job_id if cr.table2_prep else "-"
        cell_data.append([
            cell_color,
            str(cr.total_panels),
            str(len(cr.table1_panels)),
            str(len(cr.table2_panels)),
            f"{cr.forced_operator_idle}m",
            t1_prep,
            t2_prep
        ])
    
    cell_table = Table(cell_data, colWidths=[0.9*inch, 0.9*inch, 0.7*inch, 0.7*inch, 0.7*inch, 1.1*inch, 1.1*inch])
    cell_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(cell_table)
    
    # Add prep panel note if any exist
    prep_count = sum(1 for cr in result.cell_results.values() if cr.table1_prep) + \
                 sum(1 for cr in result.cell_results.values() if cr.table2_prep)
    if prep_count > 0:
        prep_note = ParagraphStyle('PrepNote', parent=styles['Normal'], fontSize=8, textColor=colors.gray, spaceAfter=4)
        elements.append(Paragraph(f"Note: {prep_count} end-of-day prep panel(s) - SETUP+LAYOUT done, ready for POUR tomorrow", prep_note))
    
    # Unscheduled jobs with extended info
    if result.unscheduled_jobs:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Unscheduled Jobs", section_style))
        
        # Add key for highlighting
        key_style = ParagraphStyle('Key', parent=styles['Normal'], fontSize=9, textColor=colors.gray)
        elements.append(Paragraph("Yellow highlight = Job is LATE (Build Date is before or on schedule date)", key_style))
        elements.append(Spacer(1, 8))
        
        unsched_headers = ["Job ID", "Req By", "Build Date", "Pri", "Class", "Reason"]
        unsched_data = [unsched_headers]
        late_rows = []  # Track which rows need yellow highlight
        
        for idx, item in enumerate(result.unscheduled_jobs[:25]):  # Limit to 25
            job = item[0]
            calc = item[1] if len(item) > 2 else None
            reason = item[-1]
            
            # Get dates and priority from calc if available
            if calc:
                req_by = str(job.req_by) if job.req_by else "-"
                build_date = str(calc.build_date) if calc.build_date else "-"
                priority = str(calc.priority)
                sched_class = calc.sched_class or "-"
                
                # Check if late (priority 0 or 1 means due today or past due)
                if calc.priority <= 1:
                    late_rows.append(idx + 1)  # +1 for header row
            else:
                req_by = str(job.req_by) if job.req_by else "-"
                build_date = "-"
                priority = "-"
                sched_class = "-"
            
            unsched_data.append([
                job.job_id[:18],
                req_by,
                build_date,
                priority,
                sched_class,
                reason[:30]
            ])
        
        # Wider Class column
        unsched_table = Table(unsched_data, colWidths=[1.4*inch, 0.8*inch, 0.8*inch, 0.4*inch, 0.6*inch, 2.5*inch])
        
        # Base table style
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c0392b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (4, -1), 'CENTER'),  # Center Priority and Class
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ]
        
        # Add yellow highlighting for late rows
        for row_idx in late_rows:
            table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#fff9c4')))
        
        unsched_table.setStyle(TableStyle(table_style))
        elements.append(unsched_table)
        
        if len(result.unscheduled_jobs) > 25:
            elements.append(Paragraph(f"... and {len(result.unscheduled_jobs) - 25} more", 
                ParagraphStyle('More', parent=styles['Normal'], fontSize=9, textColor=colors.gray)))
    
    # Footer
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Generated by Cell Scheduling Engine v1.0.0 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray)
    ))
    
    doc.build(elements)
    return output_path


def generate_debug_excel(
    result: MultiCellScheduleResult,
    job_calcs: dict,
    output_path: str | Path,
    method_name: str = "",
    variant_name: str = ""
) -> Path:
    """Generate debugging Excel file with all job data and schedule assignments.
    
    Includes:
    - Input columns (REQ_BY through ORANGE_ELIGIBLE)
    - User input columns (ON_TABLE_TODAY, JOB_QTY_REMAINING, EXPEDITE)
    - Calculated fields (SCHED_QTY through SCHED_CLASS)
    - Schedule results (BUILD_CELL, BUILD_TABLE, BUILD_SEQUENCE, etc.)
    
    Args:
        result: MultiCellScheduleResult with scheduling output.
        job_calcs: Dict of job_id -> CalculatedFields.
        output_path: Path to save Excel file.
        method_name: Scheduling method used.
        variant_name: Variant used.
    
    Returns:
        Path to generated Excel file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule Debug"
    
    # Define headers
    headers = [
        # Input columns (from DAILY_PRODUCTION_LOAD)
        "REQ_BY",
        "JOB",
        "DESCRIPTION",
        "PATTERN",
        "OPENING_SIZE",
        "WIRE_DIAMETER",
        "MOLDS",
        "MOLD_TYPE",
        "PROD_QTY",
        "EQUIVALENT",
        "ORANGE_ELIGIBLE",
        # User input columns
        "ON_TABLE_TODAY",
        "JOB_QTY_REMAINING",
        "EXPEDITE",
        # Calculated fields
        "SCHED_QTY",
        "BUILD_LOAD",
        "BUILD_DATE",
        "PRIORITY",
        "FIXTURE",
        "MOLD_DEPTH",
        "SCHED_CLASS",
        # Schedule results
        "SCHEDULED",
        "BUILD_CELL",
        "BUILD_TABLE",
        "BUILD_SEQUENCE",
        "PANELS_SCHEDULED",
        "UNSCHEDULED_REASON",
        # Additional debug columns
        "TASK_SETUP",
        "TASK_LAYOUT",
        "TASK_POUR",
        "TASK_CURE",
        "TASK_UNLOAD",
        "TOTAL_CYCLE_TIME",
    ]
    
    # Write headers
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Build lookup for scheduled jobs
    # job_id -> list of (cell_color, table_num, sequence, panels)
    scheduled_lookup = {}
    
    # Build sequence numbers from cell results
    for cell_color, cell_result in result.cell_results.items():
        # Combine T1 and T2 panels and sort by start time
        all_panels = []
        for panel in cell_result.table1_panels:
            start = panel.tasks.get("SETUP", panel.tasks.get("POUR"))
            start_time = start.start_time if start else 0
            all_panels.append((panel, 1, start_time))
        for panel in cell_result.table2_panels:
            start = panel.tasks.get("SETUP", panel.tasks.get("POUR"))
            start_time = start.start_time if start else 0
            all_panels.append((panel, 2, start_time))
        
        # Sort by start time to get sequence
        all_panels.sort(key=lambda x: x[2])
        
        for seq, (panel, table_num, _) in enumerate(all_panels, 1):
            job_id = panel.job_id
            if job_id not in scheduled_lookup:
                scheduled_lookup[job_id] = []
            scheduled_lookup[job_id].append({
                "cell": cell_color,
                "table": table_num,
                "sequence": seq,
                "panel": panel
            })
    
    # Build lookup for unscheduled jobs
    unscheduled_lookup = {}
    for item in result.unscheduled_jobs:
        job = item[0]
        reason = item[-1]
        unscheduled_lookup[job.job_id] = reason
    
    # Write job data
    row_idx = 2
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    # Get all jobs from job_calcs (which includes the Job objects)
    from .data_loader import Job
    
    # Process job assignments to get job objects
    job_objects = {}
    for assignment in result.job_assignments:
        job_objects[assignment.job.job_id] = assignment.job
    
    # Also get jobs from unscheduled
    for item in result.unscheduled_jobs:
        job = item[0]
        job_objects[job.job_id] = job
    
    # Build lookup for assigned jobs (to detect jobs assigned but not scheduled)
    assigned_lookup = {}
    for assignment in result.job_assignments:
        job_id = assignment.job.job_id
        if job_id not in assigned_lookup:
            assigned_lookup[job_id] = {
                "cell": assignment.cell_color,
                "panels_requested": 0
            }
        assigned_lookup[job_id]["panels_requested"] += assignment.panels_to_schedule
    
    # Write rows for each job
    for job_id, job in job_objects.items():
        calc = job_calcs.get(job_id)
        
        # Get schedule info
        is_scheduled = job_id in scheduled_lookup
        schedule_entries = scheduled_lookup.get(job_id, [])
        unscheduled_reason = unscheduled_lookup.get(job_id, "")
        
        # Check for jobs assigned but not scheduled (fell through the cracks)
        if not is_scheduled and not unscheduled_reason and job_id in assigned_lookup:
            assigned_info = assigned_lookup[job_id]
            unscheduled_reason = f"Assigned to {assigned_info['cell']} but no capacity"
        
        # Aggregate schedule info
        if schedule_entries:
            cells = list(set(e["cell"] for e in schedule_entries))
            tables = list(set(e["table"] for e in schedule_entries))
            sequences = [e["sequence"] for e in schedule_entries]
            panels_scheduled = len(schedule_entries)
            
            build_cell = cells[0] if len(cells) == 1 else ", ".join(cells)
            build_table = tables[0] if len(tables) == 1 else ", ".join(map(str, tables))
            build_sequence = f"{min(sequences)}-{max(sequences)}" if len(sequences) > 1 else str(sequences[0])
            
            # Get timing from first panel
            first_panel = schedule_entries[0]["panel"]
            task_setup = first_panel.tasks.get("SETUP")
            task_layout = first_panel.tasks.get("LAYOUT")
            task_pour = first_panel.tasks.get("POUR")
            task_cure = first_panel.tasks.get("CURE")
            task_unload = first_panel.tasks.get("UNLOAD")
        else:
            build_cell = ""
            build_table = ""
            build_sequence = ""
            panels_scheduled = 0
            task_setup = task_layout = task_pour = task_cure = task_unload = None
        
        # Calculate total cycle time
        if task_setup:
            total_cycle = (
                (task_setup.duration if task_setup else 0) +
                (task_layout.duration if task_layout else 0) +
                (task_pour.duration if task_pour else 0) +
                (task_cure.duration if task_cure else 0) +
                (task_unload.duration if task_unload else 0)
            )
        else:
            total_cycle = ""
        
        # Build row data
        row_data = [
            str(job.req_by) if job.req_by else "",
            job.job_id,
            job.description[:60] if job.description else "",
            str(job.pattern) if job.pattern else "",
            job.opening_size,
            job.wire_diameter,
            job.molds,
            str(job.mold_type) if job.mold_type else "",
            job.prod_qty,
            job.equivalent,
            "Yes" if job.orange_eligible else "No",
            job.on_table_today if job.on_table_today else "",
            job.job_quantity_remaining if job.job_quantity_remaining else "",
            "Yes" if job.expedite else "No",
            # Calculated fields
            calc.sched_qty if calc else "",
            round(calc.build_load, 2) if calc else "",
            str(calc.build_date) if calc else "",
            calc.priority if calc else "",
            calc.fixture_id if calc else "",
            calc.mold_depth if calc else "",
            calc.sched_class if calc else "",
            # Schedule results
            "Yes" if is_scheduled else "No",
            build_cell,
            build_table,
            build_sequence,
            panels_scheduled if panels_scheduled > 0 else "",
            unscheduled_reason,
            # Task timings
            task_setup.duration if task_setup else "",
            task_layout.duration if task_layout else "",
            task_pour.duration if task_pour else "",
            task_cure.duration if task_cure else "",
            task_unload.duration if task_unload else "",
            total_cycle if total_cycle else "",
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Highlight based on priority/schedule status
            if calc and calc.priority <= 1 and not is_scheduled:
                cell.fill = red_fill  # Late and unscheduled = red
            elif is_scheduled:
                cell.fill = green_fill  # Scheduled = green
            elif not is_scheduled and unscheduled_reason:
                cell.fill = yellow_fill  # Unscheduled = yellow
        
        row_idx += 1
    
    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        # Set reasonable widths
        if header in ["DESCRIPTION"]:
            ws.column_dimensions[col_letter].width = 40
        elif header in ["JOB", "FIXTURE", "UNSCHEDULED_REASON"]:
            ws.column_dimensions[col_letter].width = 20
        elif header in ["REQ_BY", "BUILD_DATE", "ON_TABLE_TODAY"]:
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 10
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Schedule Debug Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    
    ws_summary["A3"] = "Schedule Date:"
    ws_summary["B3"] = str(result.schedule_date)
    ws_summary["A4"] = "Shift Minutes:"
    ws_summary["B4"] = result.shift_minutes
    ws_summary["A5"] = "Method:"
    ws_summary["B5"] = f"{method_name} ({variant_name})"
    ws_summary["A6"] = "Total Panels:"
    ws_summary["B6"] = result.total_panels
    ws_summary["A7"] = "Jobs Scheduled:"
    ws_summary["B7"] = len(scheduled_lookup)
    
    # Count all unscheduled (explicit + assigned but no capacity)
    assigned_but_not_scheduled = sum(
        1 for job_id in assigned_lookup 
        if job_id not in scheduled_lookup and job_id not in unscheduled_lookup
    )
    total_unscheduled = len(unscheduled_lookup) + assigned_but_not_scheduled
    
    ws_summary["A8"] = "Jobs Unscheduled:"
    ws_summary["B8"] = total_unscheduled
    if assigned_but_not_scheduled > 0:
        ws_summary["A9"] = f"  (includes {assigned_but_not_scheduled} assigned but no capacity)"
    
    # Cell breakdown
    ws_summary["A10"] = "Cell Breakdown"
    ws_summary["A10"].font = Font(bold=True)
    ws_summary["A11"] = "Cell"
    ws_summary["B11"] = "T1 Panels"
    ws_summary["C11"] = "T2 Panels"
    ws_summary["D11"] = "Total"
    ws_summary["E11"] = "T1 Prep"
    ws_summary["F11"] = "T2 Prep"
    
    row = 12
    for cell_color, cr in sorted(result.cell_results.items()):
        ws_summary[f"A{row}"] = cell_color
        ws_summary[f"B{row}"] = len(cr.table1_panels)
        ws_summary[f"C{row}"] = len(cr.table2_panels)
        ws_summary[f"D{row}"] = cr.total_panels
        # Add prep panel info
        ws_summary[f"E{row}"] = cr.table1_prep.job_id if cr.table1_prep else ""
        ws_summary[f"F{row}"] = cr.table2_prep.job_id if cr.table2_prep else ""
        row += 1
    
    # Add prep panel summary
    prep_count = sum(1 for cr in result.cell_results.values() if cr.table1_prep) + \
                 sum(1 for cr in result.cell_results.values() if cr.table2_prep)
    if prep_count > 0:
        row += 1
        ws_summary[f"A{row}"] = f"End-of-Day Prep Panels: {prep_count}"
        ws_summary[f"A{row}"].font = Font(italic=True)
        row += 1
        ws_summary[f"A{row}"] = "(SETUP+LAYOUT done, ready for POUR tomorrow)"
        ws_summary[f"A{row}"].font = Font(italic=True, color="666666")
    
    wb.save(output_path)
    return output_path
